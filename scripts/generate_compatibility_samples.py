"""Regenerate the public, fully synthetic compatibility samples.

The generator is deterministic and intentionally keeps non-standard headers so
the files exercise the normal import, mapping and validation pipeline.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SAMPLE_DIR = ROOT / "data" / "samples"
ORDER_COUNT = 80
EVENTS_PER_ORDER = 6


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def generate_orders_csv(path: Path) -> None:
    headers = [
        "销售平台",
        "Order No",
        "创建时间",
        "promisedDeliveryTime",
        "实际送达时间",
        "orderQty",
        "已交付数量",
        "unit",
        "订单状态",
        "仓库编码",
        "物流公司",
        "目的区域",
        "无关备注",
    ]
    warehouses = ["WH-SYN-EAST-01", "WH-SYN-SOUTH-01", "WH-SYN-WEST-01"]
    carriers = ["CAR-SYN-A", "CAR-SYN-B", "CAR-SYN-C", "CAR-SYN-D"]
    regions = ["CN-SYN-EAST", "CN-SYN-NORTH", "CN-SYN-SOUTH", "CN-SYN-WEST"]
    channels = ["自营商城", "第三方平台", "直播渠道", "门店小程序"]
    rows: list[list[object]] = []
    start = datetime(2026, 7, 1, 8, 0)
    for index in range(1, ORDER_COUNT + 1):
        created = start + timedelta(days=(index - 1) % 24, minutes=index * 17)
        promised = created + timedelta(hours=72)
        quantity = 1 + index % 5
        if index % 17 == 0:
            status = "cancelled"
            actual = ""
            delivered: object = "0"
            note = "合成取消订单；用于取消率验证"
        elif index % 19 == 0:
            status = "returned"
            actual = created + timedelta(hours=62 + index % 4)
            delivered = str(quantity)
            note = "合成退回订单；用于退回率验证"
        elif index % 13 == 0:
            status = "processing"
            actual = ""
            delivered = ""
            note = "尚未交付；对应指标应保持不可计算"
        else:
            status = "已签收" if index % 2 else "妥投"
            actual = created + timedelta(hours=88 if index % 11 == 0 else 46 + index % 19)
            delivered = str(quantity - 1) if index % 9 == 0 else str(quantity)
            note = "合成晚到" if actual > promised else "合成正常履约"
        created_text = (
            created.strftime("%Y/%m/%d %H:%M")
            if index % 3 == 0
            else created.strftime("%Y-%m-%dT%H:%M:%S+08:00")
        )
        promised_text = (
            promised.strftime("%Y.%m.%d %H:%M")
            if index % 4 == 0
            else promised.strftime("%Y-%m-%d %H:%M")
        )
        actual_text = (
            actual.strftime("%Y/%m/%d %H:%M")
            if isinstance(actual, datetime) and index % 5 == 0
            else actual.strftime("%Y-%m-%dT%H:%M:%S+08:00")
            if isinstance(actual, datetime)
            else ""
        )
        rows.append(
            [
                channels[index % len(channels)],
                f"{1000000 + index:08d}",
                created_text,
                promised_text,
                actual_text,
                str(quantity),
                delivered,
                "piece",
                status,
                warehouses[index % len(warehouses)],
                carriers[index % len(carriers)] if index % 23 else "",
                regions[index % len(regions)],
                note,
            ]
        )
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(rows)


def generate_logistics_xlsx(path: Path) -> None:
    workbook = Workbook()
    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)
    orders = workbook.active
    orders.title = "订单数据"
    warehouse = workbook.create_sheet("仓库事件")
    tracking = workbook.create_sheet("物流轨迹")
    orders.append(
        [
            "仓库编码",
            "订单编号",
            "orderCreatedAt",
            "承诺送达时间",
            "actualDeliveryTime",
            "下单数量",
            "deliveredQty",
            "计量单位",
            "订单状态名称",
            "carrierCode",
            "收货地区",
            "channel",
            "可忽略备注",
        ]
    )
    warehouse.append(
        [
            "来源系统",
            "warehouseCode",
            "订单编号",
            "作业时间",
            "warehouseEventId",
            "作业状态",
            "处理数量",
            "unit",
            "扩展说明",
        ]
    )
    tracking.append(
        [
            "waybillNo",
            "轨迹状态",
            "扫描时间",
            "订单编号",
            "物流事件编号",
            "carrierCode",
            "网点代码",
            "事件序号",
            "附加备注",
        ]
    )
    warehouses = ["WH-SYN-EAST-01", "WH-SYN-SOUTH-01", "WH-SYN-WEST-01"]
    carriers = ["CAR-SYN-A", "CAR-SYN-B", "CAR-SYN-C", "CAR-SYN-D"]
    regions = ["CN-SYN-EAST", "CN-SYN-NORTH", "CN-SYN-SOUTH", "CN-SYN-WEST"]
    channels = ["自营商城", "第三方平台", "直播渠道", "门店小程序"]
    warehouse_statuses = ["仓库接单", "开始拣货", "拣货完成", "复核完成", "打包完成", "已出库"]
    tracking_statuses = ["已揽件", "始发地已发出", "运输中", "到达目的城市", "派送中", "已签收"]
    start = datetime(2026, 7, 10, 8, 0)
    for index in range(1, ORDER_COUNT + 1):
        order_id = f"XLS-{index:05d}"
        shipment_id = f"SHP-SYN-{index:05d}"
        created = start + timedelta(days=(index - 1) % 24, minutes=index * 13)
        promised = created + timedelta(hours=72)
        long_tail = index % 12 == 0
        actual = created + timedelta(hours=92 if long_tail else 50 + index % 15)
        tracking_start = created + timedelta(hours=14 if long_tail else 8)
        tracking_interval_hours = 15 if long_tail else 7 + index % 3
        actual = max(
            actual,
            tracking_start + timedelta(hours=len(tracking_statuses) * tracking_interval_hours),
        )
        quantity = 1 + index % 5
        delivered = quantity - 1 if index % 14 == 0 else quantity
        warehouse_id = warehouses[index % len(warehouses)]
        carrier_id = carriers[index % len(carriers)]
        orders.append(
            [
                warehouse_id,
                order_id,
                created,
                promised,
                actual,
                quantity,
                str(delivered),
                "piece",
                "已签收" if index % 2 else "妥投",
                carrier_id,
                regions[index % len(regions)],
                channels[index % len(channels)],
                "合成长尾" if long_tail else None,
            ]
        )
        for sequence, status in enumerate(warehouse_statuses, start=1):
            warehouse.append(
                [
                    "SYN-WMS-COMPAT",
                    warehouse_id,
                    order_id,
                    created + timedelta(hours=sequence * (2 if long_tail else 1)),
                    f"WHE-{index:05d}-{sequence:02d}",
                    status,
                    str(quantity),
                    "piece",
                    "合成作业延迟" if long_tail and sequence >= 3 else None,
                ]
            )
        for sequence, status in enumerate(tracking_statuses, start=1):
            event_time = tracking_start + timedelta(
                hours=sequence * tracking_interval_hours,
            )
            if sequence == len(tracking_statuses):
                event_time = actual
            tracking.append(
                [
                    shipment_id,
                    status,
                    event_time,
                    order_id,
                    f"TRE-{index:05d}-{sequence:02d}",
                    carrier_id,
                    f"HUB-SYN-{sequence:02d}",
                    sequence,
                    "合成长尾事件" if long_tail and sequence == 4 else None,
                ]
            )
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(path)
    temporary = path.with_suffix(".deterministic.xlsx")
    with (
        zipfile.ZipFile(path, "r") as source,
        zipfile.ZipFile(
            temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target,
    ):
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            info = zipfile.ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = original.create_system
            info.external_attr = original.external_attr
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    b'<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" '
                    b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                    b'xsi:type="dcterms:W3CDTF">'
                    b"2026-01-01T00:00:00Z</dcterms:modified>",
                    payload,
                )
            target.writestr(info, payload)
    temporary.replace(path)


def update_catalog(catalog_path: Path, csv_path: Path, xlsx_path: Path) -> None:
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    by_id = {item["sample_id"]: item for item in catalog["samples"]}
    by_id["compatibility_orders_csv"]["row_counts"] = {"orders": ORDER_COUNT}
    by_id["compatibility_orders_csv"]["sha256"] = sha256(csv_path)
    by_id["compatibility_logistics_xlsx"]["row_counts"] = {
        "orders": ORDER_COUNT,
        "warehouse_events": ORDER_COUNT * EVENTS_PER_ORDER,
        "tracking_events": ORDER_COUNT * EVENTS_PER_ORDER,
    }
    by_id["compatibility_logistics_xlsx"]["sha256"] = sha256(xlsx_path)
    catalog_path.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    csv_path = SAMPLE_DIR / "compatibility_demo_orders.csv"
    xlsx_path = SAMPLE_DIR / "compatibility_demo_logistics.xlsx"
    generate_orders_csv(csv_path)
    generate_logistics_xlsx(xlsx_path)
    update_catalog(SAMPLE_DIR / "compatibility_samples.json", csv_path, xlsx_path)
    print(f"generated orders={ORDER_COUNT}, events={ORDER_COUNT * EVENTS_PER_ORDER}")


if __name__ == "__main__":
    main()

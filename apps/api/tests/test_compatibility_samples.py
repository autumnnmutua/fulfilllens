from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
SAMPLE_DIR = ROOT / "data" / "samples"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def suggested_mapping(payload: dict[str, Any]) -> dict[str, str | None]:
    return {item["source_column"]: item["suggested_field"] for item in payload["suggestions"]}


def import_file(
    client: TestClient,
    *,
    data_type: str,
    path: Path,
    mime_type: str,
    sheet_name: str | None = None,
) -> tuple[str, dict[str, Any], dict[str, Any]]:
    uploaded = client.post(
        "/api/imports/upload",
        data={"data_type": data_type},
        files={"file": (path.name, path.read_bytes(), mime_type)},
    )
    assert uploaded.status_code == 201
    task = uploaded.json()["task"]
    parsed = client.post(
        f"/api/imports/{task['task_id']}/parse",
        json={
            **({"encoding": "utf-8-sig"} if path.suffix == ".csv" else {}),
            **({"sheet_name": sheet_name} if sheet_name is not None else {}),
        },
    )
    assert parsed.status_code == 200, parsed.text
    parsed_payload = parsed.json()
    assert parsed_payload["detected_data_type"] == data_type
    assert parsed_payload["detection_confidence"] >= 0.95
    mapping = suggested_mapping(parsed_payload)
    ignored_source_columns = sorted(source for source, target in mapping.items() if target is None)
    checked = client.put(
        f"/api/imports/{task['task_id']}/validation",
        json={
            "mapping": mapping,
            "ignored_source_columns": ignored_source_columns,
            "default_timezone": "Asia/Shanghai",
        },
    )
    assert checked.status_code == 200, checked.text
    checked_payload = checked.json()
    assert checked_payload["report"]["can_confirm"] is True, checked.text
    assert checked_payload["report"]["ignored_source_columns"] == ignored_source_columns
    assert checked_payload["report"]["unresolved_source_columns"] == []
    for row in checked_payload["normalized_preview"]:
        assert not set(ignored_source_columns) & set(row)
    confirmed = client.post(f"/api/imports/{task['task_id']}/confirm")
    assert confirmed.status_code == 200, confirmed.text
    return confirmed.json()["dataset_id"], parsed_payload, checked_payload


def test_compatibility_catalog_and_files_have_verified_content(client: TestClient) -> None:
    catalog = client.get("/api/imports/samples")

    assert catalog.status_code == 200
    payload = catalog.json()
    assert [item["sample_id"] for item in payload["samples"]] == [
        "compatibility_orders_csv",
        "compatibility_logistics_xlsx",
    ]
    for sample in payload["samples"]:
        sample_path = SAMPLE_DIR / sample["file_name"]
        downloaded = client.get(f"/api/imports/samples/{sample['sample_id']}/file")
        assert downloaded.status_code == 200
        assert downloaded.content == sample_path.read_bytes()
        assert len(downloaded.content) > 100
        assert hashlib.sha256(downloaded.content).hexdigest() == sample["sha256"]
        if sample_path.suffix == ".csv":
            with sample_path.open("r", encoding="utf-8-sig", newline="") as handle:
                assert sum(1 for _ in csv.reader(handle)) - 1 == sample["row_counts"]["orders"]
        else:
            workbook = load_workbook(sample_path, read_only=True, data_only=True)
            assert {
                "orders": workbook["订单数据"].max_row - 1,
                "warehouse_events": workbook["仓库事件"].max_row - 1,
                "tracking_events": workbook["物流轨迹"].max_row - 1,
            } == sample["row_counts"]


def test_nonstandard_csv_converts_validates_imports_and_analyzes(
    client: TestClient,
) -> None:
    dataset_id, parsed, checked = import_file(
        client,
        data_type="orders",
        path=SAMPLE_DIR / "compatibility_demo_orders.csv",
        mime_type="text/csv",
    )

    assert parsed["total_rows"] == 80
    assert parsed["preview_rows"][0]["values"]["Order No"] == "01000001"
    assert "无关备注" in parsed["unmapped_source_columns"]
    assert checked["report"]["valid_rows"] == 80
    assert checked["report"]["error_rows"] == 0
    assert checked["report"]["duplicate_keys"] == 0
    assert checked["report"]["unknown_statuses"] == 0

    metrics = client.get(
        "/api/metrics/summary",
        params={"orders_dataset_id": dataset_id},
    )
    assert metrics.status_code == 200, metrics.text
    by_code = {item["code"]: item for item in metrics.json()["metrics"]}
    assert by_code["ot_rate"]["denominator"] >= 60
    assert by_code["if_rate"]["denominator"] >= 60
    assert by_code["otif_rate"]["denominator"] >= 60


def test_multisheet_xlsx_each_sheet_converts_and_combined_metrics_reconcile(
    client: TestClient,
) -> None:
    path = SAMPLE_DIR / "compatibility_demo_logistics.xlsx"
    selections: dict[str, str] = {}
    expected = {
        "orders": ("订单数据", 80),
        "warehouse_events": ("仓库事件", 480),
        "tracking_events": ("物流轨迹", 480),
    }
    for data_type, (sheet_name, row_count) in expected.items():
        dataset_id, parsed, checked = import_file(
            client,
            data_type=data_type,
            path=path,
            mime_type=XLSX_MIME,
            sheet_name=sheet_name,
        )
        selections[data_type] = dataset_id
        assert parsed["total_rows"] == row_count
        assert checked["report"]["valid_rows"] == row_count
        assert checked["report"]["error_rows"] == 0
        assert checked["report"]["duplicate_keys"] == 0
        assert checked["report"]["unknown_statuses"] == 0
        assert any("Excel 日期单元格" in warning for warning in parsed["warnings"])

    summary = client.get(
        "/api/metrics/summary",
        params={
            "orders_dataset_id": selections["orders"],
            "warehouse_events_dataset_id": selections["warehouse_events"],
            "tracking_events_dataset_id": selections["tracking_events"],
        },
    )
    assert summary.status_code == 200, summary.text
    payload = summary.json()
    assert payload["datasets"]["orders_dataset_id"] == selections["orders"]
    assert next(item for item in payload["metrics"] if item["code"] == "order_count")["value"] == 80

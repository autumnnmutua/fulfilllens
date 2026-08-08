from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import random
import re
from dataclasses import dataclass, replace
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from typing import Literal
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo
from zoneinfo import ZoneInfo

from jsonschema import Draft202012Validator, FormatChecker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from referencing import Registry, Resource

from app.cases.models import CASE_GENERATOR_VERSION, CaseId
from app.imports.contracts import SCHEMA_DIR
from app.schemas.imports import DataType

Row = dict[str, object]
ScenarioKind = Literal["normal", "promotion", "carrier"]
TIMEZONE = ZoneInfo("Asia/Shanghai")
PRIVACY_STATEMENT = (
    "本案例完全由程序生成，不包含真实姓名、手机号、身份证、详细地址、真实快递单号、"
    "真实企业名称或任何公司内部数据。"
)

CSV_FIELDS: dict[DataType, tuple[str, ...]] = {
    DataType.ORDERS: (
        "order_id",
        "created_at",
        "promised_delivery_time",
        "actual_delivery_time",
        "ordered_quantity",
        "delivered_quantity",
        "quantity_unit",
        "order_status",
        "raw_order_status",
        "warehouse_id",
        "carrier_id",
        "destination_region",
        "sales_channel",
    ),
    DataType.WAREHOUSE_EVENTS: (
        "event_id",
        "order_id",
        "event_time",
        "event_code",
        "raw_status",
        "warehouse_id",
        "quantity",
        "quantity_unit",
        "source_system",
    ),
    DataType.TRACKING_EVENTS: (
        "tracking_event_id",
        "order_id",
        "shipment_id",
        "event_time",
        "event_code",
        "raw_status",
        "carrier_id",
        "location_code",
        "region_code",
        "exception_code",
        "sequence_number",
    ),
}


@dataclass(frozen=True)
class CaseConfig:
    case_id: CaseId
    kind: ScenarioKind
    display_name: str
    business_background: str
    seed: int
    order_count: int
    start_date: date
    day_weights: tuple[float, ...]
    injected_anomalies: tuple[str, ...]
    expected_findings: tuple[tuple[str, str, bool], ...]
    expected_metric_ranges: dict[str, tuple[float, float, str]]
    learning_objectives: tuple[str, ...]


CASE_CONFIGS: dict[CaseId, CaseConfig] = {
    CaseId.NORMAL_OPERATIONS: CaseConfig(
        case_id=CaseId.NORMAL_OPERATIONS,
        kind="normal",
        display_name="案例 A：正常运营",
        business_background="区域电商仓在非活动期保持稳定日单量，仓内与运输节点波动较小。",
        seed=20260801,
        order_count=180,
        start_date=date(2026, 6, 1),
        day_weights=(1,) * 18,
        injected_anomalies=("约 4% 订单注入可解释的随机揽收或运输延迟", "约 2% 订单部分交付"),
        expected_findings=(
            ("FL-PU-001", "少量随机订单出现出库后揽收等待", True),
            ("FL-LH-001", "足量样本下最慢约 10% 形成统计长尾", True),
        ),
        expected_metric_ranges={
            "ot_rate": (0.90, 1.00, "ratio"),
            "if_rate": (0.95, 1.00, "ratio"),
            "otif_rate": (0.87, 1.00, "ratio"),
            "fulfillment_duration_mean_hours": (35.0, 65.0, "hour"),
            "anomaly_order_rate": (0.01, 0.10, "ratio"),
        },
        learning_objectives=(
            "理解订单、仓库事件与物流轨迹的关联关系",
            "复算 OT、IF、OTIF、P50 和 P90",
            "区分正常波动、长尾和数据覆盖率",
        ),
    ),
    CaseId.PROMOTION_SURGE: CaseConfig(
        case_id=CaseId.PROMOTION_SURGE,
        kind="promotion",
        display_name="案例 B：促销爆单",
        business_background="活动开始后订单在两天内集中涌入，仓库作业能力未同步扩容。",
        seed=20260802,
        order_count=240,
        start_date=date(2026, 6, 1),
        day_weights=(1, 1, 1, 1, 1, 1, 1, 6, 5, 2, 1.5, 1.5),
        injected_anomalies=(
            "活动峰值日订单量显著上升",
            "峰值订单的接单等待、拣货、复核、打包和出库耗时同步恶化",
        ),
        expected_findings=(
            ("FL-WH-001", "仓内多个节点超过业务阈值", True),
            ("FL-WC-001", "订单量与仓内 P90 同步恶化的相关性观察", True),
        ),
        expected_metric_ranges={
            "ot_rate": (0.25, 0.85, "ratio"),
            "if_rate": (0.94, 1.00, "ratio"),
            "otif_rate": (0.20, 0.82, "ratio"),
            "fulfillment_duration_mean_hours": (45.0, 80.0, "hour"),
            "anomaly_order_rate": (0.01, 0.10, "ratio"),
        },
        learning_objectives=(
            "识别活动波峰与仓内瓶颈节点",
            "理解订单量与时长共同恶化只是相关性证据",
            "用仓内改善方案评估方向性影响",
        ),
    ),
    CaseId.CARRIER_DISRUPTION: CaseConfig(
        case_id=CaseId.CARRIER_DISRUPTION,
        kind="carrier",
        display_name="案例 C：承运商异常",
        business_background=(
            "仓内作业保持稳定，但教学承运商 CAR-SYN-SLOW 出现揽收、干线和末端长尾。"
        ),
        seed=20260803,
        order_count=180,
        start_date=date(2026, 6, 1),
        day_weights=(1,) * 18,
        injected_anomalies=(
            "CAR-SYN-SLOW 占比约 40%，揽收等待超过 12 小时",
            "CAR-SYN-SLOW 干线、目的城市停留和派送时长显著增加",
            "部分慢承运商订单带有合成运输异常状态",
        ),
        expected_findings=(
            ("FL-PU-001", "出库至揽收等待超过业务阈值", True),
            ("FL-LH-001", "承运运输时长形成明显长尾", True),
            ("FL-LM-001", "目的城市等待或派送到签收超过阈值", True),
            ("FL-CR-001", "慢承运商 OTIF/P90 相对同批其他承运商恶化", True),
        ),
        expected_metric_ranges={
            "ot_rate": (0.45, 0.75, "ratio"),
            "if_rate": (0.95, 1.00, "ratio"),
            "otif_rate": (0.42, 0.72, "ratio"),
            "fulfillment_duration_mean_hours": (75.0, 115.0, "hour"),
            "anomaly_order_rate": (0.10, 0.35, "ratio"),
        },
        learning_objectives=(
            "比较承运商 OTIF、P90 与样本量",
            "从聚合诊断下钻到揽收、干线和末端证据",
            "讨论承运商结构调整的历史重采样限制",
        ),
    ),
}


@dataclass(frozen=True)
class GeneratedCase:
    config: CaseConfig
    orders: list[Row]
    warehouse_events: list[Row]
    tracking_events: list[Row]

    def rows(self, data_type: DataType) -> list[Row]:
        if data_type == DataType.ORDERS:
            return self.orders
        if data_type == DataType.WAREHOUSE_EVENTS:
            return self.warehouse_events
        return self.tracking_events


def customized_config(case_id: CaseId, *, seed: int | None, order_count: int | None) -> CaseConfig:
    base = CASE_CONFIGS[case_id]
    return replace(
        base,
        seed=base.seed if seed is None else seed,
        order_count=base.order_count if order_count is None else order_count,
    )


def _largest_remainder(total: int, weights: tuple[float, ...]) -> list[int]:
    weight_sum = sum(weights)
    raw = [total * weight / weight_sum for weight in weights]
    values = [math.floor(value) for value in raw]
    remainder = total - sum(values)
    order = sorted(range(len(raw)), key=lambda index: (-(raw[index] - values[index]), index))
    for index in order[:remainder]:
        values[index] += 1
    return values


def _choice(rng: random.Random, weighted_values: tuple[tuple[str, float], ...]) -> str:
    threshold = rng.random() * sum(weight for _, weight in weighted_values)
    cumulative = 0.0
    for value, weight in weighted_values:
        cumulative += weight
        if threshold <= cumulative:
            return value
    return weighted_values[-1][0]


def _hours(rng: random.Random, minimum: float, maximum: float) -> float:
    return rng.uniform(minimum, maximum)


def _iso(value: datetime) -> str:
    return value.isoformat(timespec="seconds")


def _warehouse_durations(
    rng: random.Random, *, promotion_peak: bool
) -> tuple[float, float, float, float, float]:
    if promotion_peak:
        return (
            _hours(rng, 4.8, 7.0),
            _hours(rng, 2.8, 4.2),
            _hours(rng, 2.2, 3.4),
            _hours(rng, 1.8, 2.7),
            _hours(rng, 2.5, 3.8),
        )
    return (
        _hours(rng, 0.5, 0.9),
        _hours(rng, 0.7, 0.95),
        _hours(rng, 0.35, 0.55),
        _hours(rng, 0.35, 0.55),
        _hours(rng, 0.5, 0.75),
    )


def _append_warehouse_event(
    rows: list[Row],
    *,
    case_code: str,
    order_id: str,
    order_index: int,
    sequence: int,
    event_time: datetime,
    event_code: str,
    warehouse_id: str,
    quantity: int,
) -> None:
    rows.append(
        {
            "event_id": f"WHE-SYN-{case_code}-{order_index:04d}-{sequence:02d}",
            "order_id": order_id,
            "event_time": _iso(event_time),
            "event_code": event_code,
            "raw_status": event_code,
            "warehouse_id": warehouse_id,
            "quantity": quantity,
            "quantity_unit": "piece",
            "source_system": "synthetic_case_generator",
        }
    )


def _append_tracking_event(
    rows: list[Row],
    *,
    case_code: str,
    order_id: str,
    order_index: int,
    shipment_id: str,
    sequence: int,
    event_time: datetime,
    event_code: str,
    carrier_id: str,
    location_code: str,
    region_code: str,
    exception_code: str | None = None,
) -> None:
    rows.append(
        {
            "tracking_event_id": f"TRE-SYN-{case_code}-{order_index:04d}-{sequence:02d}",
            "order_id": order_id,
            "shipment_id": shipment_id,
            "event_time": _iso(event_time),
            "event_code": event_code,
            "raw_status": event_code,
            "carrier_id": carrier_id,
            "location_code": location_code,
            "region_code": region_code,
            "exception_code": exception_code,
            "sequence_number": sequence,
        }
    )


def generate_case(config: CaseConfig) -> GeneratedCase:
    if config.order_count < 30:
        raise ValueError("教学案例至少需要 30 个订单，才能支持分位数和分组讲解。")
    rng = random.Random(config.seed)
    daily_counts = _largest_remainder(config.order_count, config.day_weights)
    orders: list[Row] = []
    warehouse_events: list[Row] = []
    tracking_events: list[Row] = []
    case_code = {
        CaseId.NORMAL_OPERATIONS: "A",
        CaseId.PROMOTION_SURGE: "B",
        CaseId.CARRIER_DISRUPTION: "C",
    }[config.case_id]
    order_index = 0

    for day_index, daily_count in enumerate(daily_counts):
        current_date = config.start_date + timedelta(days=day_index)
        for position in range(daily_count):
            order_index += 1
            promotion_peak = config.kind == "promotion" and day_index in {7, 8}
            hour = 9 + ((position * 11 + order_index * 3) % 11)
            if promotion_peak:
                hour = 9 + ((position * 5 + order_index) % 6)
            minute = (position * 17 + order_index * 7) % 60
            created_at = datetime.combine(current_date, time(hour, minute), tzinfo=TIMEZONE)
            order_id = f"ORD-SYN-{case_code}-{order_index:04d}"
            shipment_id = f"SHP-SYN-{case_code}-{order_index:04d}"
            warehouse_id = _choice(rng, (("WH-SYN-EAST", 0.72), ("WH-SYN-WEST", 0.28)))
            if config.kind == "carrier":
                carrier_id = _choice(
                    rng,
                    (("CAR-SYN-SLOW", 0.40), ("CAR-SYN-FAST", 0.35), ("CAR-SYN-STD", 0.25)),
                )
            else:
                carrier_id = _choice(
                    rng,
                    (("CAR-SYN-FAST", 0.45), ("CAR-SYN-STD", 0.35), ("CAR-SYN-ECON", 0.20)),
                )
            region = _choice(
                rng,
                (("CN-SD-QD", 0.30), ("CN-JS-NJ", 0.27), ("CN-ZJ-HZ", 0.25), ("CN-HB-WH", 0.18)),
            )
            sales_channel = _choice(
                rng,
                (
                    ("synthetic_marketplace", 0.55),
                    ("synthetic_store", 0.30),
                    ("synthetic_classroom", 0.15),
                ),
            )
            quantity = rng.randint(1, 5)

            received = created_at + timedelta(hours=_hours(rng, 0.1, 0.4))
            order_to_pick, picking, pick_to_qc, quality_check, packing = _warehouse_durations(
                rng, promotion_peak=promotion_peak
            )
            picking_started = received + timedelta(hours=order_to_pick)
            picking_completed = picking_started + timedelta(hours=picking)
            quality_started = picking_completed + timedelta(hours=pick_to_qc)
            quality_completed = quality_started + timedelta(hours=quality_check)
            packing_started = quality_completed + timedelta(hours=_hours(rng, 0.05, 0.2))
            packing_completed = packing_started + timedelta(hours=packing)
            ready_to_ship = packing_completed + timedelta(hours=_hours(rng, 0.15, 0.5))
            shipped = ready_to_ship + timedelta(hours=_hours(rng, 0.05, 0.2))

            warehouse_points = (
                (received, "order_received"),
                (picking_started, "picking_started"),
                (picking_completed, "picking_completed"),
                (quality_started, "quality_check_started"),
                (quality_completed, "quality_check_completed"),
                (packing_started, "packing_started"),
                (packing_completed, "packing_completed"),
                (ready_to_ship, "ready_to_ship"),
                (shipped, "shipped_from_warehouse"),
            )
            for sequence, (event_time, event_code) in enumerate(warehouse_points, start=1):
                _append_warehouse_event(
                    warehouse_events,
                    case_code=case_code,
                    order_id=order_id,
                    order_index=order_index,
                    sequence=sequence,
                    event_time=event_time,
                    event_code=event_code,
                    warehouse_id=warehouse_id,
                    quantity=quantity,
                )

            is_slow_carrier = config.kind == "carrier" and carrier_id == "CAR-SYN-SLOW"
            random_delay = config.kind == "normal" and rng.random() < 0.04
            if is_slow_carrier:
                pickup_wait = _hours(rng, 18, 30)
                first_leg = _hours(rng, 38, 52)
                hub_dwell = _hours(rng, 5, 10)
                second_leg = _hours(rng, 38, 55)
                destination_dwell = _hours(rng, 28, 42)
                delivery_attempt = _hours(rng, 14, 22)
            else:
                pickup_wait = _hours(rng, 2.5, 8.0)
                first_leg = _hours(rng, 8, 15)
                hub_dwell = _hours(rng, 1, 4)
                second_leg = _hours(rng, 8, 16)
                destination_dwell = _hours(rng, 2, 9)
                delivery_attempt = _hours(rng, 1, 6)
                if random_delay:
                    if order_index % 2:
                        pickup_wait += 14
                    else:
                        second_leg += 50

            shipment_created = ready_to_ship
            picked_up = ready_to_ship + timedelta(hours=pickup_wait)
            origin_departed = picked_up + timedelta(hours=_hours(rng, 0.5, 1.5))
            in_transit = origin_departed + timedelta(hours=_hours(rng, 0.5, 1.5))
            arrived_hub = picked_up + timedelta(hours=first_leg)
            departed_hub = arrived_hub + timedelta(hours=hub_dwell)
            arrived_destination = departed_hub + timedelta(hours=second_leg)
            out_for_delivery = arrived_destination + timedelta(hours=destination_dwell)
            delivered = out_for_delivery + timedelta(hours=delivery_attempt)
            tracking_points: list[tuple[datetime, str, str, str | None]] = [
                (shipment_created, "shipment_created", f"ORIGIN-{warehouse_id}", None),
                (picked_up, "carrier_picked_up", f"ORIGIN-{warehouse_id}", None),
                (origin_departed, "origin_departed", f"ORIGIN-{warehouse_id}", None),
                (in_transit, "in_transit", "LINEHAUL-SYN-01", None),
            ]
            if is_slow_carrier and order_index % 2 == 0:
                exception_time = in_transit + (arrived_hub - in_transit) / 2
                tracking_points.append(
                    (exception_time, "exception", "LINEHAUL-SYN-01", "SYNTHETIC_DELAY")
                )
            tracking_points.extend(
                [
                    (arrived_hub, "arrived_at_hub", "HUB-SYN-01", None),
                    (departed_hub, "departed_hub", "HUB-SYN-01", None),
                    (arrived_destination, "arrived_at_destination_city", f"DEST-{region}", None),
                    (out_for_delivery, "out_for_delivery", f"SITE-{region}", None),
                    (delivered, "delivered", f"SITE-{region}", None),
                ]
            )
            tracking_points.sort(key=lambda item: item[0])
            for sequence, (event_time, event_code, location_code, exception_code) in enumerate(
                tracking_points, start=1
            ):
                _append_tracking_event(
                    tracking_events,
                    case_code=case_code,
                    order_id=order_id,
                    order_index=order_index,
                    shipment_id=shipment_id,
                    sequence=sequence,
                    event_time=event_time,
                    event_code=event_code,
                    carrier_id=carrier_id,
                    location_code=location_code,
                    region_code=region,
                    exception_code=exception_code,
                )

            promised_hours = 72.0
            if config.kind == "promotion":
                promised_hours = 60.0
            promised = created_at + timedelta(hours=promised_hours)
            partial_probability = 0.02 if config.kind != "promotion" else 0.03
            delivered_quantity = quantity if rng.random() >= partial_probability else quantity - 1
            orders.append(
                {
                    "order_id": order_id,
                    "created_at": _iso(created_at),
                    "promised_delivery_time": _iso(promised),
                    "actual_delivery_time": _iso(delivered),
                    "ordered_quantity": quantity,
                    "delivered_quantity": delivered_quantity,
                    "quantity_unit": "piece",
                    "order_status": "delivered",
                    "raw_order_status": "delivered",
                    "warehouse_id": warehouse_id,
                    "carrier_id": carrier_id,
                    "destination_region": region,
                    "sales_channel": sales_channel,
                }
            )

    return GeneratedCase(
        config=config,
        orders=orders,
        warehouse_events=warehouse_events,
        tracking_events=tracking_events,
    )


def _validator(data_type: DataType) -> Draft202012Validator:
    schema_file = {
        DataType.ORDERS: "order.schema.json",
        DataType.WAREHOUSE_EVENTS: "warehouse_event.schema.json",
        DataType.TRACKING_EVENTS: "tracking_event.schema.json",
    }[data_type]
    schema = json.loads((SCHEMA_DIR / schema_file).read_text(encoding="utf-8"))
    status_schema = json.loads(
        (SCHEMA_DIR / "status_codes.schema.json").read_text(encoding="utf-8")
    )
    registry = Registry().with_resource(status_schema["$id"], Resource.from_contents(status_schema))
    return Draft202012Validator(schema, registry=registry, format_checker=FormatChecker())


def validate_generated_case(generated: GeneratedCase) -> list[str]:
    errors: list[str] = []
    order_ids = {str(row["order_id"]) for row in generated.orders}
    for data_type in DataType:
        validator = _validator(data_type)
        rows = generated.rows(data_type)
        primary = {
            DataType.ORDERS: "order_id",
            DataType.WAREHOUSE_EVENTS: "event_id",
            DataType.TRACKING_EVENTS: "tracking_event_id",
        }[data_type]
        identifiers = [str(row[primary]) for row in rows]
        if len(identifiers) != len(set(identifiers)):
            errors.append(f"{data_type.value}: primary key is not unique")
        for index, row in enumerate(rows, start=2):
            for error in validator.iter_errors(row):
                errors.append(f"{data_type.value}:{index}:{error.message}")
            if data_type != DataType.ORDERS and str(row["order_id"]) not in order_ids:
                errors.append(f"{data_type.value}:{index}: orphan order_id")
    for rows in (generated.warehouse_events, generated.tracking_events):
        previous: dict[tuple[str, str], datetime] = {}
        for row in rows:
            key = (str(row["order_id"]), str(row.get("shipment_id") or ""))
            timestamp = datetime.fromisoformat(str(row["event_time"]))
            if key in previous and timestamp < previous[key]:
                errors.append(f"{key[0]}: event timestamps are out of order")
            previous[key] = timestamp
    return errors


def _logical_fingerprint(generated: GeneratedCase) -> str:
    payload = {
        "generator_version": CASE_GENERATOR_VERSION,
        "seed": generated.config.seed,
        "orders": generated.orders,
        "warehouse_events": generated.warehouse_events,
        "tracking_events": generated.tracking_events,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def metadata_payload(generated: GeneratedCase) -> dict[str, object]:
    config = generated.config
    end_date = config.start_date + timedelta(days=len(config.day_weights) - 1)
    return {
        "case_id": config.case_id.value,
        "display_name": config.display_name,
        "business_background": config.business_background,
        "generator_version": CASE_GENERATOR_VERSION,
        "seed": config.seed,
        "timezone": "Asia/Shanghai",
        "order_count": len(generated.orders),
        "date_range": {"start": config.start_date.isoformat(), "end": end_date.isoformat()},
        "row_counts": {
            "orders": len(generated.orders),
            "warehouse_events": len(generated.warehouse_events),
            "tracking_events": len(generated.tracking_events),
        },
        "injected_anomalies": list(config.injected_anomalies),
        "expected_findings": [
            {"rule_id": rule_id, "description": description, "required": required}
            for rule_id, description, required in config.expected_findings
        ],
        "expected_metric_ranges": {
            code: {"minimum": values[0], "maximum": values[1], "unit": values[2]}
            for code, values in config.expected_metric_ranges.items()
        },
        "learning_objectives": list(config.learning_objectives),
        "privacy_statement": PRIVACY_STATEMENT,
        "content_fingerprint": _logical_fingerprint(generated),
        "files": [],
    }


def _csv_bytes(rows: list[Row], fields: tuple[str, ...]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=fields, lineterminator="\n", extrasaction="raise")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def _xlsx_bytes(generated: GeneratedCase) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is None:
        raise RuntimeError("新建工作簿缺少默认工作表")
    workbook.remove(default_sheet)
    workbook.properties.creator = "FulfillLens CN synthetic case generator"
    fixed_time = datetime(2026, 1, 1, tzinfo=UTC).replace(tzinfo=None)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    for data_type in DataType:
        sheet = workbook.create_sheet(data_type.value)
        fields = CSV_FIELDS[data_type]
        sheet.append(list(fields))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1677FF")
        for row in generated.rows(data_type):
            sheet.append([row.get(field) for field in fields])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    raw = io.BytesIO()
    workbook.save(raw)
    source = io.BytesIO(raw.getvalue())
    deterministic = io.BytesIO()
    with (
        ZipFile(source, "r") as original,
        ZipFile(deterministic, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for name in sorted(original.namelist()):
            existing = original.getinfo(name)
            info = ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = existing.external_attr
            info.create_system = existing.create_system
            content = original.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    (
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                        b"2026-01-01T00:00:00Z</dcterms:modified>"
                    ),
                    content,
                )
            target.writestr(info, content, compress_type=ZIP_DEFLATED, compresslevel=9)
    return deterministic.getvalue()


def write_case_artifacts(
    generated: GeneratedCase, output_root: Path, *, include_xlsx: bool = True
) -> Path:
    errors = validate_generated_case(generated)
    if errors:
        raise ValueError("合成案例未通过校验：" + "; ".join(errors[:10]))
    case_dir = output_root / generated.config.case_id.value
    case_dir.mkdir(parents=True, exist_ok=True)
    files: dict[str, tuple[bytes, str]] = {}
    for data_type in DataType:
        name = f"{data_type.value}.csv"
        files[name] = (
            _csv_bytes(generated.rows(data_type), CSV_FIELDS[data_type]),
            "text/csv; charset=utf-8",
        )
    if include_xlsx:
        files["case.xlsx"] = (
            _xlsx_bytes(generated),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    for name, (content, _) in files.items():
        (case_dir / name).write_bytes(content)
    metadata = metadata_payload(generated)
    metadata["files"] = [
        {"name": name, "media_type": media_type, "size_bytes": len(content)}
        for name, (content, media_type) in sorted(files.items())
    ]
    metadata_bytes = (
        json.dumps(metadata, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")
    (case_dir / "metadata.json").write_bytes(metadata_bytes)
    return case_dir

from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.read_only import ReadOnlyCell

from app.core.config import Settings
from app.core.errors import AppError
from app.imports.security import inspect_xlsx_archive
from app.schemas.imports import SheetInfo

SUPPORTED_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk")
PREVIEW_ROWS = 20
MAX_CSV_FIELD_BYTES = 1024 * 1024


@dataclass(frozen=True)
class ParseIssue:
    code: str
    message: str
    row_number: int | None
    source_column: str | None
    raw_value: str | None
    suggestion: str
    severity: str = "error"


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    values: dict[str, Any]


@dataclass(frozen=True)
class ParsedTable:
    headers: list[str]
    rows: list[ParsedRow]
    sheet_name: str | None
    issues: list[ParseIssue] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def detect_csv_encoding(path: Path) -> tuple[str | None, list[str]]:
    sample = path.read_bytes()[: 64 * 1024]
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig", []
    try:
        sample.decode("utf-8", errors="strict")
    except UnicodeDecodeError:
        return None, ["gb18030", "gbk"]
    return "utf-8", []


def validate_selected_encoding(encoding: str) -> str:
    normalized = encoding.strip().casefold().replace("_", "-")
    aliases = {
        "utf8": "utf-8",
        "utf8-sig": "utf-8-sig",
        "cp936": "gbk",
        "gb2312": "gbk",
    }
    normalized = aliases.get(normalized, normalized)
    if normalized not in SUPPORTED_ENCODINGS:
        raise AppError(
            code="UNSUPPORTED_CSV_ENCODING",
            message="CSV 编码仅支持 UTF-8、UTF-8 BOM、GB18030 和 GBK。",
            status_code=422,
        )
    return normalized


def normalize_header(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value)).strip()


def validate_headers(headers: list[str]) -> None:
    if not headers or all(not header for header in headers):
        raise AppError(
            code="MISSING_HEADER_ROW",
            message="文件第一行必须包含字段名。",
            status_code=422,
        )
    if any(not header for header in headers):
        raise AppError(
            code="EMPTY_COLUMN_NAME",
            message="表头包含空列名，请在原文件中补充名称。",
            status_code=422,
        )
    normalized = [header.casefold() for header in headers]
    duplicates = sorted(header for header in set(normalized) if normalized.count(header) > 1)
    if duplicates:
        raise AppError(
            code="DUPLICATE_COLUMN_NAMES",
            message="表头包含重复列名，必须先改为唯一名称。",
            status_code=422,
        )


def parse_csv(path: Path, *, encoding: str, settings: Settings) -> ParsedTable:
    selected_encoding = validate_selected_encoding(encoding)
    try:
        text = path.read_text(encoding=selected_encoding, errors="strict")
    except UnicodeDecodeError as error:
        raise AppError(
            code="CSV_DECODING_FAILED",
            message=f"文件无法按 {selected_encoding} 严格解码，请选择其他编码。",
            status_code=422,
        ) from error

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel

    previous_limit = csv.field_size_limit()
    csv.field_size_limit(MAX_CSV_FIELD_BYTES)
    try:
        rows = csv.reader(text.splitlines(), dialect=dialect)
        raw_headers = next(rows, None)
        if raw_headers is None:
            raise AppError(
                code="EMPTY_FILE",
                message="CSV 文件为空。",
                status_code=400,
            )
        headers = [normalize_header(value) for value in raw_headers]
        if len(headers) > settings.max_import_columns:
            raise AppError(
                code="TOO_MANY_COLUMNS",
                message="文件列数超过安全上限。",
                status_code=413,
            )
        validate_headers(headers)

        parsed_rows: list[ParsedRow] = []
        issues: list[ParseIssue] = []
        for row_number, values in enumerate(rows, start=2):
            if row_number - 1 > settings.max_import_rows:
                raise AppError(
                    code="TOO_MANY_ROWS",
                    message="文件行数超过当前导入上限。",
                    status_code=413,
                )
            if not values or all(value == "" for value in values):
                continue
            if len(values) != len(headers):
                issues.append(
                    ParseIssue(
                        code="COLUMN_COUNT_MISMATCH",
                        message="该行列数与表头不一致。",
                        row_number=row_number,
                        source_column=None,
                        raw_value=None,
                        suggestion="检查分隔符、引号和缺失单元格。",
                    )
                )
            normalized_values = (values + [""] * len(headers))[: len(headers)]
            parsed_rows.append(
                ParsedRow(
                    row_number=row_number,
                    values=dict(zip(headers, normalized_values, strict=True)),
                )
            )
    except csv.Error as error:
        raise AppError(
            code="CSV_PARSING_FAILED",
            message="CSV 结构无法解析，请检查引号、分隔符和换行。",
            status_code=422,
        ) from error
    finally:
        csv.field_size_limit(previous_limit)

    if not parsed_rows:
        raise AppError(
            code="NO_DATA_ROWS",
            message="文件仅包含表头或空行。",
            status_code=422,
        )
    return ParsedTable(headers=headers, rows=parsed_rows, sheet_name=None, issues=issues)


def list_xlsx_sheets(path: Path, settings: Settings) -> list[SheetInfo]:
    inspect_xlsx_archive(
        path,
        max_entries=settings.max_xlsx_entries,
        max_uncompressed_bytes=settings.max_xlsx_uncompressed_bytes,
    )
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise AppError(
            code="XLSX_PARSING_FAILED",
            message="XLSX 工作簿无法安全读取。",
            status_code=422,
        ) from error
    try:
        return [
            SheetInfo(name=worksheet.title, state=worksheet.sheet_state)
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def serialize_excel_value(
    cell: Cell | MergedCell | ReadOnlyCell,
) -> tuple[Any, ParseIssue | None]:
    value = cell.value
    if cell.data_type == "f":
        return (
            None,
            ParseIssue(
                code="FORMULA_CELL_IGNORED",
                message="公式单元格未执行，当前值按缺失处理。",
                row_number=cell.row,
                source_column=None,
                raw_value="[公式已忽略]",
                suggestion="在原工作簿中复制并粘贴为值后重新导入。",
            ),
        )
    if cell.data_type == "e":
        return (
            None,
            ParseIssue(
                code="EXCEL_ERROR_CELL",
                message="Excel 错误单元格无法解析。",
                row_number=cell.row,
                source_column=None,
                raw_value="[Excel 错误值]",
                suggestion="修复工作簿中的错误值并粘贴为普通值。",
            ),
        )
    if isinstance(value, datetime):
        return value.isoformat(), None
    if isinstance(value, date):
        return datetime.combine(value, time.min).isoformat(), None
    return value, None


def parse_xlsx(
    path: Path,
    *,
    sheet_name: str,
    settings: Settings,
) -> ParsedTable:
    inspect_xlsx_archive(
        path,
        max_entries=settings.max_xlsx_entries,
        max_uncompressed_bytes=settings.max_xlsx_uncompressed_bytes,
    )
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise AppError(
            code="XLSX_PARSING_FAILED",
            message="XLSX 工作簿无法安全读取。",
            status_code=422,
        ) from error

    try:
        if sheet_name not in workbook.sheetnames:
            raise AppError(
                code="WORKSHEET_NOT_FOUND",
                message="选择的工作表不存在。",
                status_code=422,
            )
        worksheet = workbook[sheet_name]
        if worksheet.max_row > settings.max_import_rows + 1:
            raise AppError(
                code="TOO_MANY_ROWS",
                message="工作表声明的行数超过当前导入上限。",
                status_code=413,
            )
        if worksheet.max_column > settings.max_import_columns:
            raise AppError(
                code="TOO_MANY_COLUMNS",
                message="工作表列数超过安全上限。",
                status_code=413,
            )

        row_iterator = worksheet.iter_rows()
        header_cells = next(row_iterator, None)
        if header_cells is None:
            raise AppError(
                code="EMPTY_WORKSHEET",
                message="选择的工作表为空。",
                status_code=422,
            )
        header_values = [serialize_excel_value(cell)[0] for cell in header_cells]
        headers = [normalize_header(value) if value is not None else "" for value in header_values]
        validate_headers(headers)

        parsed_rows: list[ParsedRow] = []
        issues: list[ParseIssue] = []
        for cells in row_iterator:
            row_number = (cells[0].row or len(parsed_rows) + 2) if cells else len(parsed_rows) + 2
            values: list[Any] = []
            row_issues: list[ParseIssue] = []
            for index, cell in enumerate(cells[: len(headers)]):
                value, issue = serialize_excel_value(cell)
                values.append(value)
                if issue is not None:
                    row_issues.append(
                        ParseIssue(
                            code=issue.code,
                            message=issue.message,
                            row_number=row_number,
                            source_column=headers[index],
                            raw_value=issue.raw_value,
                            suggestion=issue.suggestion,
                        )
                    )
            values.extend([None] * (len(headers) - len(values)))
            if all(value in {None, ""} for value in values):
                continue
            parsed_rows.append(
                ParsedRow(
                    row_number=row_number,
                    values=dict(zip(headers, values, strict=True)),
                )
            )
            issues.extend(row_issues)

        if not parsed_rows:
            raise AppError(
                code="NO_DATA_ROWS",
                message="工作表仅包含表头或空行。",
                status_code=422,
            )
        warnings = (
            ["当前工作表为隐藏状态；请确认它确实是要导入的数据。"]
            if worksheet.sheet_state != "visible"
            else []
        )
        return ParsedTable(
            headers=headers,
            rows=parsed_rows,
            sheet_name=sheet_name,
            issues=issues,
            warnings=warnings,
        )
    finally:
        workbook.close()
